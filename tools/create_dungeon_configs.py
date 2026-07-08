#!/usr/bin/env python3
"""一次性脚本：生成 dungeon.xlsx 与 exp_level.xlsx"""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = ROOT / "tools" / "config"

dungeon_headers = [
    ("id", "int", "", "", "主键"),
    ("name_key", "string", "text", "", "名称"),
    ("desc_key", "string", "text", "", "描述"),
    ("icon", "string", "", "", "图标路径"),
    ("min_player_level", "int", "", "", "最低等级"),
    ("recommended_level", "int", "", "", "推荐等级"),
    ("difficulty_tier", "int", "", "", "难度层级"),
    ("total_waves", "int", "", "", "总波次"),
    ("boss_wave", "int", "", "", "Boss波次"),
    ("base_enemy_level", "int", "", "", "怪物基础等级"),
    ("wave_time_base", "float", "", "", "波次基础秒数"),
    ("wave_time_increment", "float", "", "", "波次递增秒数"),
    ("max_enemies_per_wave", "int", "", "", "每波最大敌人数"),
    ("spawn_interval_base", "float", "", "", "刷怪间隔基值"),
    ("spawn_interval_min", "float", "", "", "刷怪间隔最小值"),
    ("gold_multiplier", "float", "", "", "金币倍率"),
    ("exp_multiplier", "float", "", "", "经验倍率"),
    ("loot_quality_bonus", "int", "", "", "掉落品质加成"),
    ("boss_config_id", "int", "", "", "Boss配置ID"),
    ("scene_path", "string", "", "", "关卡场景"),
    ("required_cleared_dungeon_id", "int", "", "", "需通关的前置副本"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "dungeon"
for c, h in enumerate(dungeon_headers, 1):
    ws.cell(1, c, h[0])
    ws.cell(2, c, h[1])
    ws.cell(3, c, h[2])
    ws.cell(4, c, h[3])
    ws.cell(5, c, h[4])

rows = [
    [1, "幽暗洞穴", "阴森的洞穴深处潜伏着危险的生物", "", 1, 3, 1, 6, 6, 1, 25, 2, 20, 2.0, 0.6, 1.0, 1.0, 0, 101, "res://scenes/gameplay/levels/arena.tscn", 0],
    [2, "幽暗森林", "茂密的森林中隐藏着更强大的敌人", "", 5, 8, 2, 7, 7, 3, 28, 2, 22, 1.8, 0.5, 1.2, 1.2, 1, 102, "res://scenes/gameplay/levels/arena.tscn", 1],
    [3, "遗忘城堡", "古老城堡的废墟中盘踞着亡灵", "", 10, 13, 3, 8, 8, 5, 30, 2, 24, 1.6, 0.5, 1.5, 1.5, 2, 103, "res://scenes/gameplay/levels/arena.tscn", 0],
]
for ri, row in enumerate(rows, 6):
    for ci, val in enumerate(row, 1):
        ws.cell(ri, ci, val)
wb.save(CONFIG_DIR / "dungeon.xlsx")

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "exp_level"
exp_headers = [
    ("id", "int", "", "", "等级"),
    ("exp_required", "int", "", "", "升级所需经验"),
    ("reward_hp", "int", "", "", "升级奖励生命"),
    ("reward_damage", "int", "", "", "升级奖励伤害"),
    ("reward_speed", "int", "", "", "升级奖励移速"),
    ("reward_gold", "int", "", "", "升级奖励金币"),
]
for c, h in enumerate(exp_headers, 1):
    ws2.cell(1, c, h[0])
    ws2.cell(2, c, h[1])
    ws2.cell(5, c, h[4])
for lvl in range(1, 21):
    exp_req = 50 * lvl * lvl + 50 * lvl
    r = lvl + 5
    ws2.cell(r, 1, lvl)
    ws2.cell(r, 2, exp_req)
    ws2.cell(r, 3, 5 + lvl)
    ws2.cell(r, 4, 1)
    ws2.cell(r, 5, 5)
    ws2.cell(r, 6, 50)
wb2.save(CONFIG_DIR / "exp_level.xlsx")
print("Created dungeon.xlsx and exp_level.xlsx")
