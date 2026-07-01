#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""一次性脚本：按装备系统文档补齐配置表字段与底材行。"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = ROOT / "tools" / "config"


def first_sheet(wb):
    return next(w for n in wb.sheetnames if not n.startswith("!") for w in [wb[n]])


def col_index(ws, name):
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == name:
            return c
    return None


def ensure_column(ws, name, type_str, tag="", mode="", desc=""):
    idx = col_index(ws, name)
    if idx is None:
        idx = ws.max_column + 1
        ws.cell(1, idx, name)
        ws.cell(2, idx, type_str)
        ws.cell(3, idx, tag)
        ws.cell(4, idx, mode)
        ws.cell(5, idx, desc)
    return idx


def patch_quality():
    path = CONFIG_DIR / "quality.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = first_sheet(wb)

    affix_col = ensure_column(ws, "affix_count", "int", "", "", "随机词缀数量")
    aspect_col = ensure_column(ws, "has_aspect", "bool", "", "", "是否携带威能")

    affix_counts = {1: 0, 2: 1, 3: 4, 4: 4, 5: 0}
    has_aspects = {1: False, 2: False, 3: False, 4: True, 5: True}

    for r in range(6, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        if qid is None:
            continue
        qid = int(qid)
        ws.cell(r, affix_col, affix_counts.get(qid, 0))
        ws.cell(r, aspect_col, has_aspects.get(qid, False))

    wb.save(path)
    print(f"patched {path.name}")


def patch_item():
    path = CONFIG_DIR / "item.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = first_sheet(wb)

    aspect_col = ensure_column(ws, "aspect_id", "string", "", "", "预设威能 ID（传奇底材）")
    base_col = ensure_column(ws, "is_drop_base", "bool", "", "", "是否参与随机掉落底材池")

    id_col = col_index(ws, "id")
    name_col = col_index(ws, "name_key")

    for r in range(6, ws.max_row + 1):
        item_id = ws.cell(r, id_col).value
        if item_id is None:
            continue
        item_id = int(item_id)
        name_key = ws.cell(r, name_col).value or ""

        ws.cell(r, base_col, item_id < 2000)

        if item_id >= 2000 and name_key.startswith("item.legend."):
            suffix = name_key.replace("item.legend.", "")
            ws.cell(r, aspect_col, f"aspect.{suffix}")

    # 防具白底材（随机掉落池）
    new_rows = [
        (1041, "item.helm_leather", "item.helm_leather.desc", "res://assets/icons/equit/armor_leather.png",
         2, 1, 2, 0, 0, 0, 0, 0, 5, 0, 20, True, ""),
        (1042, "item.chest_leather", "item.chest_leather.desc", "res://assets/icons/equit/armor_leather.png",
         3, 1, 3, 0, 0, 0, 0, 0, 8, 0, 25, True, ""),
        (1043, "item.boots_leather", "item.boots_leather.desc", "res://assets/icons/equit/armor_leather_boots.png",
         4, 1, 4, 0, 0, 0, 0, 0, 3, 0, 22, True, ""),
        (1044, "item.amulet_simple", "item.amulet_simple.desc", "res://assets/icons/equit/armor_cloth.png",
         5, 1, 5, 0, 5, 0, 0, 0, 0, 0, 18, True, ""),
    ]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    start_row = ws.max_row + 1
    for offset, row in enumerate(new_rows):
        r = start_row + offset
        data = {
            "id": row[0],
            "desc": None,
            "name_key": row[1],
            "desc_key": row[2],
            "icon": row[3],
            "type": row[4],
            "rarity": row[5],
            "slot_type": row[6],
            "level_req": row[7],
            "stat_hp": row[8],
            "stat_damage": row[9],
            "stat_speed": row[10],
            "stat_crit": row[11],
            "stat_armor": row[12],
            "stack_limit": row[13],
            "sell_price": row[14],
            "is_drop_base": row[15],
            "aspect_id": row[16],
        }
        for c, header in enumerate(headers, start=1):
            if header in data:
                ws.cell(r, c, data[header])

    wb.save(path)
    print(f"patched {path.name}")


def create_affix():
    path = CONFIG_DIR / "affix.xlsx"
    if path.exists():
        print(f"skip existing {path.name}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "affix"

    headers = [
        ("id", "string", "", "", "词条 ID"),
        ("numeric_type", "int", "", "", "NumericType 枚举值"),
        ("modifier_type", "int", "", "", "ModifierType 枚举值 1=固定 2=百分比"),
        ("min_per_level", "float", "", "", "每级最小系数"),
        ("max_per_level", "float", "", "", "每级最大系数"),
        ("slot_mask", "int", "comma", "", "可出现的装备槽 slot_type 列表"),
    ]
    for c, (name, typ, tag, mode, desc) in enumerate(headers, start=1):
        ws.cell(1, c, name)
        ws.cell(2, c, typ)
        ws.cell(3, c, tag)
        ws.cell(4, c, mode)
        ws.cell(5, c, desc)

    rows = [
        ("affix.vitality", 100, 1, 2, 5, "2,3,4,5"),
        ("affix.might", 102, 2, 0.01, 0.03, "1,5"),
        ("affix.swiftness", 104, 2, 0.005, 0.015, "4,5"),
        ("affix.precision", 110, 2, 0.01, 0.04, "1,2,5"),
        ("affix.fortitude", 111, 1, 1, 3, "2,3,4"),
    ]
    for r, row in enumerate(rows, start=6):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)

    wb.save(path)
    print(f"created {path.name}")


if __name__ == "__main__":
    patch_quality()
    patch_item()
    create_affix()
