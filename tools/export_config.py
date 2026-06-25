#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配置表导出工具 v1.0

功能:
  1. 读取 tools/config/ 下所有 .xlsx 文件
  2. 导出 JSON 配置到 assets/config/

模板格式说明 (以 item.xlsx 为例):
  Row 1: 字段名 (field_name)
  Row 2: 字段类型 (int | string | float | bool)
  Row 3: 字段标签 (可选: ignore, text, comma, table, json)
  Row 4: 字段模式 (可选: !mode, client)
  Row 5: 注释说明
  Row 6+: 数据行

标签含义:
  ignore  - 跳过该列，不导出
  text    - 本地化 key 字段
  comma   - 逗号分隔的列表，导出为数组
  table   - 表格格式 ({{k1,v1},{k2,v2}}), 保持原样
  json    - JSON 格式字符串，按 string 原样导出（不解析为对象）

使用方式:
  python tools/export_config.py
"""

import json
import os
import re
import sys
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl 库，正在安装...")
    import subprocess
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "openpyxl", "-q"]
    )
    import openpyxl

# 设置 stdout 编码为 utf-8
if sys.stdout.encoding and sys.stdout.encoding.upper() != "UTF-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# -- 路径配置 ---------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent  # E:/_project/hope
XLSX_DIR = ROOT / "tools" / "config"
OUTPUT_JSON_DIR = ROOT / "assets" / "config"


def parse_value(raw, type_str: str, tag: str):
    """
    根据类型和标签解析单元格值。
    返回适合 JSON 序列化的 Python 对象。
    """
    if raw is None:
        return None

    # json 标签: JSON 格式字符串，按 string 原样导出
    if tag == "json":
        if isinstance(raw, float) and raw == int(raw):
            return str(int(raw))
        return str(raw)

    # comma 标签: 逗号分割为数组
    if tag == "comma":
        if isinstance(raw, (int, float)):
            return [int(raw)]
        s = str(raw).strip()
        if not s:
            return []
        parts = [p.strip() for p in s.split(",") if p.strip()]
        result = []
        for p in parts:
            try:
                result.append(int(p))
            except ValueError:
                try:
                    result.append(float(p))
                except ValueError:
                    result.append(p)
        return result

    # table 标签: 解析 {{k1,v1},{k2,v2}} 格式为 JSON 数组 [[k1,v1],[k2,v2]]
    if tag == "table":
        s = str(raw).strip() if raw is not None else ""
        if not s or s in ("{}", "{{}}"):
            return []
        # 去掉外层 {{ 和 }}
        inner = s.strip()
        if inner.startswith("{{") and inner.endswith("}}"):
            inner = inner[2:-2]
        if not inner:
            return []
        # 按 },{ 分割成多对
        pairs = re.split(r'\}\s*,\s*\{', inner)
        result = []
        for pair in pairs:
            # 去掉残留的 { }
            pair = pair.strip().strip("{").strip("}")
            # 按逗号分割（最多分割一次，因为只有两个元素）
            parts = [p.strip().strip('"').strip("'") for p in pair.split(",", 1)]
            # 尝试转为 int 提升可用性
            typed = []
            for p in parts:
                try:
                    typed.append(int(p))
                except ValueError:
                    typed.append(p)
            result.append(typed)
        return result

    # 按类型解析
    type_str = type_str.lower()
    if type_str == "int":
        if isinstance(raw, str):
            raw = raw.strip()
            if raw == "":
                return 0
            try:
                return int(float(raw))
            except ValueError:
                return 0
        return int(raw)

    if type_str == "float":
        if isinstance(raw, str):
            raw = raw.strip()
            if raw == "":
                return 0.0
        return float(raw)

    if type_str == "bool":
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, str):
            return raw.strip().lower() in ("1", "true", "yes")
        return bool(raw)

    # string 类型
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    return str(raw)


def parse_sheet(ws):
    """
    解析一个符合模板格式的工作表。
    """
    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 6:
        return None

    raw_names = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    raw_types = [ws.cell(2, c).value for c in range(1, max_col + 1)]
    raw_tags = [ws.cell(3, c).value for c in range(1, max_col + 1)]
    raw_modes = [ws.cell(4, c).value for c in range(1, max_col + 1)]

    fields = []
    for col_idx in range(max_col):
        name = raw_names[col_idx]
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        type_str = str(raw_types[col_idx]).strip().lower() if raw_types[col_idx] else "string"
        tag = str(raw_tags[col_idx]).strip().lower() if raw_tags[col_idx] else ""
        mode = str(raw_modes[col_idx]).strip().lower() if raw_modes[col_idx] else ""

        if tag == "ignore":
            continue

        fields.append((name, type_str, tag, mode))

    if not fields:
        return None

    raw_col_indices = []
    for ci in range(max_col):
        raw_name = raw_names[ci]
        if raw_name is None or str(raw_name).strip() == "":
            continue
        raw_tag = raw_tags[ci]
        if raw_tag is not None and str(raw_tag).strip().lower() == "ignore":
            continue
        raw_col_indices.append(ci)

    data_rows = []
    for row_idx in range(6, max_row + 1):
        row_data = []
        is_empty = True
        for ci in raw_col_indices:
            cell_val = ws.cell(row_idx, ci + 1).value
            if cell_val is not None:
                is_empty = False
            fi = raw_col_indices.index(ci)
            f_name, f_type, f_tag, f_mode = fields[fi]
            parsed = parse_value(cell_val, f_type, f_tag)
            row_data.append(parsed)

        if not is_empty:
            data_rows.append(row_data)

    return {"fields": fields, "rows": data_rows}


def export_json(sheet_name: str, data: dict):
    """导出 JSON 文件"""
    fields = data["fields"]
    rows = data["rows"]

    if not fields or not rows:
        return None

    key_field = fields[0][0]

    result_dict = OrderedDict()
    for row in rows:
        item = OrderedDict()
        for i, (f_name, f_type, f_tag, f_mode) in enumerate(fields):
            if i < len(row):
                item[f_name] = row[i]
        item_key = str(row[0]) if row and row[0] is not None else ""
        if item_key:
            result_dict[item_key] = item

    output = {
        "_key": key_field,
        "_count": len(result_dict),
        "_dict": result_dict,
    }

    OUTPUT_JSON_DIR.mkdir(parents=True, exist_ok=True)

    out_path = OUTPUT_JSON_DIR / f"{sheet_name}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  [OK] JSON -> {out_path.name}  ({len(result_dict)} 条记录)")
    return output


def process_xlsx(file_path: Path):
    """处理单个 xlsx 文件"""
    print(f"\n 处理: {file_path.name}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    total_sheets = 0
    total_rows = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("!"):
            print(f"  跳过辅助表: {sheet_name}")
            continue

        ws = wb[sheet_name]
        result = parse_sheet(ws)
        if result is None or not result["rows"]:
            print(f"  空表或无数据: {sheet_name}")
            continue

        total_sheets += 1
        total_rows += len(result["rows"])

        export_json(sheet_name, result)

    if total_sheets == 0:
        print(f"  [WARN] 无有效工作表")
    else:
        print(f"  [INFO] 共 {total_sheets} 个工作表, {total_rows} 条数据")

    wb.close()


def main():
    print("=" * 50)
    print("  配置表导出工具 v1.0")
    print("=" * 50)
    print(f"  输入目录: {XLSX_DIR}")
    print(f"  JSON 输出: {OUTPUT_JSON_DIR}")
    print("=" * 50)

    if not XLSX_DIR.exists():
        print(f"[ERR] 输入目录不存在: {XLSX_DIR}")
        sys.exit(1)

    xlsx_files = sorted(f for f in XLSX_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
    if not xlsx_files:
        print(f"[WARN] 在 {XLSX_DIR} 下未找到 .xlsx 文件")
        return

    print(f"\n 找到 {len(xlsx_files)} 个 xlsx 文件:")
    for f in xlsx_files:
        print(f"    - {f.name}")

    for f in xlsx_files:
        process_xlsx(f)

    print("\n" + "=" * 50)
    print("  [OK] 全部导出完成")
    print("=" * 50)


if __name__ == "__main__":
    main()
