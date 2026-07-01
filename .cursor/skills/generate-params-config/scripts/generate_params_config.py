#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate scripts/config/ParamsConfig.cs from tools/config/params.xlsx."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[4]
DEFAULT_XLSX = ROOT / "tools" / "config" / "params.xlsx"
DEFAULT_OUT = ROOT / "scripts" / "config" / "ParamsConfig.cs"

CSHARP_RESERVED = {
    "abstract", "as", "base", "bool", "break", "byte", "case", "catch", "char",
    "checked", "class", "const", "continue", "decimal", "default", "delegate",
    "do", "double", "else", "enum", "event", "explicit", "extern", "false",
    "finally", "fixed", "float", "for", "foreach", "goto", "if", "implicit",
    "in", "int", "interface", "internal", "is", "lock", "long", "namespace",
    "new", "null", "object", "operator", "out", "override", "params", "private",
    "protected", "public", "readonly", "ref", "return", "sbyte", "sealed",
    "short", "sizeof", "stackalloc", "static", "string", "struct", "switch",
    "this", "throw", "true", "try", "typeof", "uint", "ulong", "unchecked",
    "unsafe", "ushort", "using", "virtual", "void", "volatile", "while",
}


def snake_to_pascal(name: str) -> str:
    parts = re.split(r"[_\s]+", name.strip())
    return "".join(p[:1].upper() + p[1:] for p in parts if p)


def to_property_name(name_key: str) -> str:
    safe = re.sub(r"[^0-9a-zA-Z_]+", "_", name_key.strip())
    prop = snake_to_pascal(safe)
    if not prop:
        raise ValueError(f"invalid name_key: {name_key!r}")
    if prop[0].isdigit():
        prop = "_" + prop
    if prop.lower() in CSHARP_RESERVED:
        prop = "_" + prop
    return prop


def is_filled(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def infer_field(val_num, val_str, val_color):
    candidates = []
    if is_filled(val_num):
        candidates.append(("float", float(val_num)))
    if is_filled(val_str):
        candidates.append(("string", str(val_str).strip()))
    if is_filled(val_color):
        candidates.append(("color", str(val_color).strip()))
    if not candidates:
        return None
    if len(candidates) > 1:
        print(
            f"[WARN] multiple val_* filled for row; using priority "
            f"val_num > val_str > val_color",
            file=sys.stderr,
        )
    return candidates[0]


def parse_color(raw: str, row_num: int) -> tuple[float, float, float, float | None]:
    s = raw.strip()
    if s.startswith("#"):
        hex_body = s[1:]
        if len(hex_body) == 6:
            r = int(hex_body[0:2], 16) / 255.0
            g = int(hex_body[2:4], 16) / 255.0
            b = int(hex_body[4:6], 16) / 255.0
            return r, g, b, None
        if len(hex_body) == 8:
            a = int(hex_body[0:2], 16) / 255.0
            r = int(hex_body[2:4], 16) / 255.0
            g = int(hex_body[4:6], 16) / 255.0
            b = int(hex_body[6:8], 16) / 255.0
            return r, g, b, a
        raise ValueError(f"row {row_num}: invalid hex color {s!r}")

    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) not in (3, 4):
        raise ValueError(f"row {row_num}: invalid color {s!r}")

    nums = [float(p) for p in parts]
    if max(nums) > 1.0:
        nums = [n / 255.0 for n in nums]
    if len(nums) == 3:
        return nums[0], nums[1], nums[2], None
    return nums[0], nums[1], nums[2], nums[3]


def escape_csharp_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def format_float(value: float) -> str:
    text = f"{value:.9g}"
    if "." not in text and "e" not in text.lower():
        text += "f"
    elif not text.endswith("f"):
        text += "f"
    return text


def emit_member(name_key: str, kind: str, value, comment: str | None, row_num: int) -> str:
    prop = to_property_name(name_key)
    lines = ["    /// <summary>", f"    /// {name_key}"]
    if comment:
        lines.append(f"    /// {comment}")
    lines.append("    /// </summary>")

    if kind == "float":
        lines.append(f"    public const float {prop} = {format_float(value)};")
    elif kind == "string":
        lines.append(f'    public const string {prop} = "{escape_csharp_string(value)}";')
    else:
        r, g, b, a = parse_color(value, row_num)
        if a is None:
            lines.append(
                f"    public static readonly Color {prop} = new({format_float(r)}, "
                f"{format_float(g)}, {format_float(b)});"
            )
        else:
            lines.append(
                f"    public static readonly Color {prop} = new({format_float(r)}, "
                f"{format_float(g)}, {format_float(b)}, {format_float(a)});"
            )
    return "\n".join(lines)


def read_params(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = next(w for n in wb.sheetnames if not n.startswith("!") for w in [wb[n]])

    headers = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(1, col).value
        if name:
            headers[str(name).strip()] = col

    required = {"id", "name_key", "val_num", "val_str", "val_color"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"missing columns: {', '.join(sorted(missing))}")

    comments = {}
    for col_name, col_idx in headers.items():
        desc = ws.cell(5, col_idx).value
        if desc:
            comments[col_name] = str(desc).strip()

    rows = []
    for row_idx in range(6, ws.max_row + 1):
        name_key = ws.cell(row_idx, headers["name_key"]).value
        if name_key is None or str(name_key).strip() == "":
            continue
        name_key = str(name_key).strip()
        val_num = ws.cell(row_idx, headers["val_num"]).value
        val_str = ws.cell(row_idx, headers["val_str"]).value
        val_color = ws.cell(row_idx, headers["val_color"]).value
        inferred = infer_field(val_num, val_str, val_color)
        if inferred is None:
            print(f"[WARN] row {row_idx} ({name_key}): no val_* set, skipped", file=sys.stderr)
            continue
        kind, value = inferred
        row_comment = None
        if "desc" in headers:
            desc_val = ws.cell(row_idx, headers["desc"]).value
            if desc_val is not None and str(desc_val).strip():
                row_comment = str(desc_val).strip()
        if row_comment is None:
            row_comment = comments.get("name_key")
        rows.append(
            {
                "row": row_idx,
                "name_key": name_key,
                "kind": kind,
                "value": value,
                "comment": row_comment,
            }
        )
    return rows


def generate_cs(rows) -> str:
    parts = [
        "using Godot;",
        "",
        "namespace Hope.Config;",
        "",
        "/// <summary>",
        "/// 全局参数 - 对应 tools/config/params.xlsx（自动生成，请勿手改）。",
        "/// </summary>",
        "public static partial class ParamsConfig",
        "{",
    ]

    seen = set()
    for row in rows:
        prop = to_property_name(row["name_key"])
        if prop in seen:
            raise ValueError(f"duplicate name_key -> {prop} (row {row['row']})")
        seen.add(prop)
        parts.append(emit_member(row["name_key"], row["kind"], row["value"], row["comment"], row["row"]))
        parts.append("")

    if len(parts) > 8 and parts[-1] == "":
        parts.pop()
    parts.append("}")
    parts.append("")
    return "\n".join(parts)


def main():
    parser = argparse.ArgumentParser(description="Generate ParamsConfig.cs from params.xlsx")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"[ERR] xlsx not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    rows = read_params(args.xlsx)
    content = generate_cs(rows)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(content, encoding="utf-8")
    print(f"[OK] {args.out} ({len(rows)} params)")


if __name__ == "__main__":
    main()
